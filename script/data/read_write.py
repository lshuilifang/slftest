'''
@File:slftest.py
@DateTime:2021/11/22
@Author:sunlufan
@Desc:
'''

import openpyxl

class ReadWrite:
    def __init__(self,file,sheet):
        self.file=file
        self.sheet=sheet
        self.wb=openpyxl.load_workbook(self.file)
        self.table=self.wb[self.sheet]
        self.wb.active
        self.maxrow=self.table.max_row
        self.maxcol=self.table.max_column

    def Read(self):
        list2=[]
        for i in range(2,self.maxrow+1):
            list1=[]
            for j in range(1,self.maxcol+1):
                values=self.table.cell(i,j).value
                list1.append(values)
            list2.append(list1)
        self.wb.close()
        return list2

    def Write(self,*arg):
        for i in range(1,len(arg)+1):
            self.table.cell(self.maxrow+1,i).value=arg[i-1]
        self.wb.save(self.file)
        self.wb.close()


if __name__=="__main__":
    file=r"E:\slftest1\script\data\test.xlsx"
    sheet="Sheet1"
    doc1=ReadWrite(file,sheet)
    print(doc1.Read())
    # doc1.Write("sunlufan","123456")


